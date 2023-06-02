import openpyxl #installar openpyxl
import xlrd # installar xlrd
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

def info():
    '''
    Estructura de uso basica:
        import openpyxl

        # Cargar el archivo Excel existente
        workbook = openpyxl.load_workbook('nombre_del_archivo.xlsx')

        # Seleccionar una hoja de trabajo
        hoja = workbook['nombre_de_la_hoja']

        # Leer el contenido de una celda
        valor = hoja['A1'].value
        print(f"El valor de la celda A1 es: {valor}")

        # Escribir en una celda
        hoja['B1'] = '¡Hola desde Python!'

        # Guardar el archivo
        workbook.save('nombre_del_archivo.xlsx')

        # Cerrar el archivo
        workbook.close()
    '''
    return None

print(info.__doc__)

# Leer el archivo .xls con xlrd
archivo_xls = xlrd.open_workbook(r'PysimpleGUI\Proyectos\mantto_feeder\MF-64_plantilla.XLS')
hoja_xls = archivo_xls.sheet_by_index(0)

# Crear un nuevo archivo .xlsx con openpyxl
archivo_xlsx = openpyxl.Workbook()
hoja_xlsx = archivo_xlsx.active

# Copiar los datos de .xls a .xlsx
for fila in range(hoja_xls.nrows):
    for columna in range(hoja_xls.ncols):
        valor_celda = hoja_xls.cell_value(fila, columna)
        hoja_xlsx.cell(row=fila+1, column=columna+1).value = valor_celda

# Guardar el archivo .xlsx
archivo_xlsx.save('MF-64_23760055.xlsx')
archivo_xlsx.close()


