import openpyxl
from openpyxl import workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_template():
    '''
    Funcionamiento basico:
        - Cargar el archivo Excel existente(MF-64_plantilla.XLS)
        - lo convertimos a XLSX(MF-64_plantilla.xlsx)
        ** version mas actual de archivos excel
        -recabamos toda la info de los parametros y generamos una plantilla para guardar como reporte
    parametros(no definidos en funcion aun):
        -Nombre del tecnico
        -ID feeder
        -Tipo de feeder
        -Fecha de mantenimiento
        -Color de semana
    '''
    nombre_excel = r'PysimpleGUI\Proyectos\mantto_feeder\data\MF-64_plantilla.xlsx'
    
    try:
        # Intentar cargar el archivo existente
        workbook = openpyxl.load_workbook(nombre_excel)
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        workbook = Workbook()
        workbook.save(nombre_excel)
    
    #*****************Obtener la hoja de trabajo*******************
    hoja = workbook.active
    font = Font(name='Arial', size=10, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    
    #*******************fusion de celdas*************************
    
    #fecha
    hoja.merge_cells("B12:C12")
    #tenico
    hoja.merge_cells("F12:G12")
    #id feeder
    hoja.merge_cells("B14:C14")
    #color
    hoja.merge_cells("F14:G14")
    
    #******************** Rellenar celdas ***********************
    hoja["B12"].value = "30/05/2023"
    hoja["F12"].value = "Cristian Echevarria"
    hoja["B14"].value = "23760055"
    hoja["F14"].value = "ROJO"
    hoja["H18"].value = "OK"
    
    #*********************Aplicar estilos************************
    hoja["B12"].font = font
    hoja["F12"].font = font
    hoja["B14"].font = font
    hoja["F14"].font = font
    hoja["H18"].font = font
    
    hoja["B12"].alignment = alignment
    hoja["F12"].alignment = alignment
    hoja["B14"].alignment = alignment
    hoja["F14"].alignment = alignment
    hoja["H18"].alignment = alignment
    
    #guardar workbook
    ID_feeder = hoja["B14"].value
    nuevo_nombre = f"PysimpleGUI\Proyectos\mantto_feeder\MF-{ID_feeder}.xlsx"
    workbook.save(nuevo_nombre)
    workbook.close()
    
    
    return create_template.__doc__
 
print(create_template())
create_template()