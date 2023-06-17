import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import time


# ************************************* MANIPULACION DE ARCHIVOS EXCEL *************************************
def info():
    '''
    Estructura de uso basica:
        import openpyxl

        # Cargar el archivo Excel existente
        workbook = openpyxl.load_workbook('nombre_del_archivo.xlsx')

        # Seleccionar una hoja de trabajo
        hoja = workbook['nombre_de_la_hoja']

        # Leer el contenido de una celda
        valor = hoja['A1'].value
        print(f"El valor de la celda A1 es: {valor}")

        # Escribir en una celda
        hoja['B1'] = '¡Hola desde Python!'

        # Guardar el archivo
        workbook.save('nombre_del_archivo.xlsx')

        # Cerrar el archivo
        workbook.close()
    '''
    return None

print(info.__doc__)


# Nombre del archivo Excel
nombre_excel = r'PysimpleGUI\Proyectos\mantto_feeder\manto_feeder.xlsx'

try:
    # Intentar cargar el archivo existente
    workbook = openpyxl.load_workbook(nombre_excel)
except FileNotFoundError:
    # Si el archivo no existe, crear uno nuevo
    workbook = openpyxl.Workbook()
    workbook.save(nombre_excel)

# Obtener la hoja de trabajo
hoja = workbook.active

#rellenar o cambiar el valor de una una celda
hoja["A1"].value = "Nombre"
#o de la sig.forma
hoja["B1"].value = "Apellido"
hoja["A2"].value = "Cristian"
hoja["B2"].value = "Echevarria"

#fusionar dos celdas
hoja.merge_cells("C1:D1")
hoja["C1"].value = "datos fusionados"

#fusion de celdas en diagonal para cuadros de texto
hoja.merge_cells("A10:F17") 

#Cambiar el estilo de fuente y tamaño de letra
font = Font(name='Arial', size=20, bold=True)
hoja["A1"].font = font
hoja["B1"].font = font

#Centrar datos en una celda
alignment = Alignment(horizontal="center", vertical="center")
hoja["C1"].alignment = alignment

#definir el tamaño de la celda
hoja.column_dimensions["A"].width = 20

#ajuste automatico de celda segun el texto
hoja.column_dimensions["B"].auto_size = True

#************************************* Rellenar un rango de columnas *************************************

# Definir el rango de columnas que deseas llenar
columna_inicio = 'J'
columna_fin = 'V'
fila_inicio = 1
fila_fin = 1

# Llenar el rango de columnas con la palabra "OK" (RELLENAR RANGO O CUADRO COMO EL DE OBSERVACION)
for columna in range(ord(columna_inicio), ord(columna_fin) + 1):
    for fila in range(fila_inicio, fila_fin + 1):
        celda = chr(columna) + str(fila)
        hoja[celda] = "OK"

#************************** Rellenar un rango de columnas *************************************
# Detenerse si encuentra una celda con datos
columna_inicio = 'J'
columna_fin = 'V'
fila_inicio = 3
fila_fin = 3

# Bandera para indicar si se encuentra una celda con datos
encontrado_datos = False

# Llenar el rango de columnas con la palabra "OK"
for columna in range(ord(columna_inicio), ord(columna_fin) + 1):
    for fila in range(fila_inicio, fila_fin + 1):
        celda = chr(columna) + str(fila)
        if hoja[celda].value is None:
            hoja[celda] = "OK"
        else:
            encontrado_datos = True
            break
    if encontrado_datos:
        break

# ************************** Buscar un valor o valores en una columna o fila ********************************

# Especificar la columna en la que se realizará la búsqueda
columna = hoja['A']  # Suponiendo que deseas buscar en la columna A

# Valor a buscar
valor_deseado = None # valor a buscar igual a None para poder hacer un condicional que busque cualquier valor diferente de cero es decir que nos muestre todos los valores en la columna

# Iterar sobre las celdas de la columna
for celda in columna:
    if celda.value != valor_deseado:
        fila = celda.row
        valor = str(celda.value).replace("00:00:00","")
        print("Valor encontrado en la fila:",fila,"con un valor de:",valor.replace("-","/"))
        

#*************************  Buscar valores en rango de columnas************************************
columna_inicio = 'J'
columna_fin = 'Z'
fila_inicio = 5
fila_fin = 5

# Bandera para indicar si se encuentra una celda con datos
encontrado_datos = False

#busqueda de valores
for columna in range(ord(columna_inicio), ord(columna_fin) + 1):
    for fila in range(fila_inicio, fila_fin + 1):
        celda = chr(columna) + str(fila)
        if hoja[celda].value is None:
            None
        else:
            encontrado_datos = True
            valor = str(hoja[celda].value).replace("00:00:00","")
            print("Se encontró el valor en la fila:",fila,"con un valor de :", valor.replace("-","/"))


#**************************** rellenar la interseccion de una fila con una columna ************************
# Definir fila y columna
fila = 6
columna = 'J'

# Calcular la intersección
columna_interseccion = chr(ord(columna))
fila_interseccion = str(int(fila))

# Obtener la referencia de la celda de intersección
celda_interseccion = hoja[columna_interseccion + fila_interseccion]

# Asignar un valor a la celda de intersección
celda_interseccion.value = "OK"

#***************************** GUARDAR Y CERRAR ARCHIVO ****************************     
#nuevo_nombre = "MF-64_23760055.xls"
# Guardar y cerrar el archivo
workbook.save("PysimpleGUI\Proyectos\mantto_feeder\MF-64_23760055.xlsx")
workbook.close()

